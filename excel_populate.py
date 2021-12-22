from datetime import date
from openpyxl import *
import re
from selenium import webdriver
import time


wb = load_workbook("C:/Users/Mike G/Documents/bet/wgucapfin.xlsx")
ws = wb["Sheet1"]                                                       #load excel sheet

driver = webdriver.Firefox()                #creates initial webdriver with firefox
action = webdriver.ActionChains(driver)



invalid_list = ['Final', 'City', 'OT', 'Golden', 'New', 'LA', 'San']

def insert_home_data(data_list, row_index):         #inserts teams into home team column
    home_team = data_list[0]
    home_score = data_list[1]
    insert_data(home_team, row_index, 2)
    insert_data(home_score, row_index, 3)

def insert_away_data(data_list, row_index):         # inserts teams into away team column
    away_team = data_list[2]
    away_score = data_list[3]
    insert_data(away_team, row_index, 4)
    insert_data(away_score, row_index, 5)


def insert_teams(new_row_index):     #recursive algorithm to enter previous scores into excel data sheet
    print('INSERT TEAMSSSSSSSSSSSSSSSSSSSS')
    today = date.today()
    day = today.strftime("%d")
    month = today.strftime("%m")
    year = today.strftime("%yyyy")
    yesterday = int(day) - 1
    driver.get("https://www.basketball-reference.com/boxscores/?month=12&day=19&year=2021")         #loads yesterdays games webpage

    teams = driver.find_elements_by_class_name("teams")
    row_index = new_row_index
    print('ROW INDEX ' + str(row_index))
    insert_date(row_index)          # inserts data at row index
    for team in teams:
        words = team.text.split()
        if(words):
            cleaned_data = clean_data(words, row_index)
            insert_home_data(cleaned_data, row_index)
            insert_away_data(cleaned_data, row_index)
            row_index+=1
            if row_index == 2350:
                break
    row_index+=2
    click_previous_day(row_index)

def endtest():
    print('end')

def click_previous_day(prev_row_index):
    prev_link = driver.find_element_by_xpath("/html/body/div[2]/div[3]/div[1]/a[1]")
    prev_link.click()
    insert_teams(prev_row_index)


def insert_date(row_index):
    date = driver.find_element_by_xpath("/html/body/div[2]/div[3]/h1")
    fixed_date = date.text[20:len(date.text)]                               #get dates from page and formats properly
    insert_data(fixed_date, row_index, 1)


def insert_data(data, row_index, column_index):
    print(data)
    print(row_index)
    print(column_index)
    cell_insert = ws.cell(row_index, column_index)                          # inserts data into particular index on excel
    cell_insert.value = data
    print(str(data) + " INSERTED INTO ROW: " + str(row_index) + " COLUMN: " + str(column_index))
    wb.save("C:/Users/Mike G/Documents/bet/wgucapfin.xlsx")

def get_data(row_index, col_index):         #returns excel cell data given row and column
    c = ws.cell(row_index, col_index)
    value = c.value
    return value

def clean_data(data_list, row_index):
    if 'OT' in data_list:
        insert_data('OT', row_index, 7)
    for invalid in invalid_list:
        if invalid in data_list:
            data_list.remove(invalid)
    return data_list


nba_teams = ['76ers', "Blazers", 'Bucks', 'Bulls', 'Cavaliers', 'Celtics', 'Clippers', 'Grizzlies', 'Hawks', 'Heat', 'Hornets', 'Jazz', 'Kings', 'Knicks', 'Lakers', 'Magic', 'Mavericks', 'Nets', 'Nuggets', 'Pacers', 'Pelicans', 'Pistons', 'Raptors', 'Rockets', 'Spurs', 'Suns', 'Thunder', 'Timberwolves', 'Trail Blazers', "Trailblazers", 'Warriors', 'Wizards']


def team_dict(team):
    team_dict = {
        'Antonio' : 'Spurs',
        'Atlanta':'Hawks',
        'Atlanta Hawks' : 'Hawks',
        'Blazers':'Trailblazers',
        'Boston':'Celtics',
        'Brooklyn': 'Nets',
        'Bucks':'Bucks',
        'Bulls':'Bulls',
        'Cavs':'Cavaliers',
        'Cavaliers':'Cavaliers',
        'Celtics': 'Celtics',
        'Celitcs': 'Celtics',
        'Charlotte': 'Hornets',
        'Chicago': 'Bulls',
        'Clippers' : 'Clippers',
        'Cleveland':'Cavaliers',
        'Dallas':'Mavericks',
        'Denver':'Nuggets',
        'Detroit':'Pistons',
        'ers':'76ers',
        'Golden' : 'Warriors',
        'Golden State Warriors':'Warriors',
        'Grizzlies': 'Grizzlies',
        'Indiana':'Pacers',
        'Indiana Pacers' : 'Pacers',
        'Jazz':'Jazz',
        'Hawks':'Hawks',
        'Heat':'Heat',
        'Hornets':'Hornets',
        'Houston':'Rockets',
        'Kings':'Kings',
        'Knicks':'Knicks',
        'Lakers': 'Lakers',
        'Lakers ' : 'Lakers',
        'LA Clippers':'Clippers',
        'Los Angeles Clippers':'Clippers',
        'Los Angeles Lakers':'Lakers',
        'Magic':'Magic',
        'Mavericks':'Mavericks',
        'Memphis':'Grizzlies',
        'Miami':'Heat',
        'Milwaukee':'Bucks',
        'Milwaukee Bucks':'Bucks',
        'Minnesota':'Timberwolves',
        'Nets':'Nets',
        'New York Knicks':'Knicks',
        'Nuggets':'Nuggets',
        'Nuggets ' : 'Nuggets',
        'Okla' : 'Thunder',
        'OKC': 'Thunder',
        'Orlando':'Magic',
        'Orleans':'Pelicans',
        'Pacers':'Pacers',
        'Pelicans':'Pelicans',
        'Pelicans ' : 'Pelicans',
        'Phoenix':'Suns',
        'Phoenix Suns' : 'Suns',
        'Philadelphia': '76ers',
        'Pistons':'Pistons',
        'Portland':'Trailblazers',
        'Raptors':'Raptors',
        'Rockets':'Rockets',
        'Sacramento':'Kings',
        'Sacramento Kings' : 'Kings',
        'Spurs':'Spurs',
        'Thunder':'Thunder',
        'Timerwolves':'Timberwolves',
        'Toronto':'Raptors',
        'Trail Blazers':'Trail Blazers',
        'Trailblazers':'Trailblazers',
        'City':'Thunder',
        'State':'Warriors',
        'Suns':'Suns',
        'Timberwolves':'Timberwolves',
        'Trail':'Trailblazers',
        'Utah':'Jazz',
        'Utah Jazz' : 'Jazz',
        'Washington':'Wizards',
        'Washington Wizards':'Wizards',
        'Warriors':'Warriors',
        'Wizards':'Wizards',
        'York':'Knicks',
        '76ers':'76ers'
    }
    return team_dict[team]


def get_last_year_records():            #gets last years records for format into team objects
    driver.get("https://www.basketball-reference.com/leagues/NBA_2020_standings.html")
    rows = driver.find_element_by_xpath('//*[@id="expanded_standings"]')
    valid_data = []
    data = rows.text
    data_list = data.split()

    del data_list[0:32]             #clears unnecessary data from list
    for stat in data_list:
        if stat in nba_teams:
            valid_index = data_list.index(stat)

            score_index = valid_index+1
            score = data_list[score_index]
            valid_data.append(stat)
            valid_data.append(score)

    for i in range(0, len(valid_data), 2):              #inserts team name and wins into excel sheet
        team_name = valid_data[i]
        if team_name == 'Blazers':
            team_name = 'Trailblazers'
        team_record = valid_data[i+1]
        team_wins = team_record[0:2]
        team_wins = int(team_wins)
        new_index = i/2+1
        new_index = int(new_index)
        insert_data(team_name, new_index, 10)
        insert_data(team_wins, new_index, 11)

def get_last_ten():
    driver.get("https://www.teamrankings.com/nba/ranking/last-10-games-by-other")
    rows = driver.find_elements_by_tag_name("tr")
    row_index = 35
    for row in rows:
        stat_list = row.text.split()
        if (stat_list) and stat_list[0].isdigit():
            clean_data(stat_list, 2)
            team_loc = stat_list[1]
            team_name = team_dict(team_loc)
            last_10 = stat_list[2]
            last_10_wins = last_10[1:2]
            rating = stat_list[3]
            correct_index = team_match(team_name)
            print(last_10)
            #insert_data(team_name, correct_index, 10)
            insert_data(int(last_10_wins), correct_index, 12)
            #insert_data(rating, row_index, 12)
            row_index+=1
            print('---------')

def get_team_elos():                    #gets team elo ratings
    driver.get("https://projects.fivethirtyeight.com/2022-nba-predictions/")
    count = 0
    rows = driver.find_elements_by_tag_name("tr")
    print(rows)
    for row in rows:
        if any(x in row.text for x in nba_teams):
            stats = row.text.split()
            team_elo = stats[0]
            team_loc = stats[1]                             #splits strings up and matches them in dictionary to format team names

            team_format = "".join(re.split("[^a-zA-Z]*", team_loc))
            team_name = team_dict(team_format)
            correct_index = team_match(team_name)               #adds elo rating to correct index
            insert_data(int(team_elo), correct_index, 13)



def team_match(data):                   #matches team and returns index of row

    for i in range(1,31):
        team_name = get_data(i, 10)
        if data in team_name:
            print('MATCHED ' + str(data) + ' WITH ' + str(team_name) + ' AT ' + str(i))
            return i

def populate():
    #insert_teams(2)
    get_last_year_records()
    get_last_ten()
    get_team_elos()